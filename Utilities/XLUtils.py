import openpyxl

def getRowCount(file , sheet):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet]
    return sheet.max_row

def readXLData(file , sheet , rowno , colno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet]
    return sheet.cell(row=rowno , column=colno).value